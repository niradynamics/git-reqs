import unittest
import os
import sys
import shutil
from openpyxl import load_workbook, Workbook
sys.path.append("../git_reqs")
import requirementmodule
from exporttools import convert_to_xls

class TestGitReqs(unittest.TestCase):
    def test_init(self):
        if os.path.exists("test_init"):
            shutil.rmtree("test_init", ignore_errors=True)

        os.system('../git-reqs init --module test_init --module_prefix TEST1 --root_module')
        os.system('cd test_init && ../../git-reqs init --module test_init2 --module_prefix TEST2 --root_module')
        os.system('cd test_init/test_init2 && ../../../git-reqs init --module test_init3 --module_prefix TEST3')
        os.system('cd test_init/test_init2 && ../../../git-reqs init --module test_init4 --module_prefix TEST4 --req_numbering_format numbers')

        req_module = requirementmodule.requirement_module("./test_init")

        self.assertEqual(req_module.config['module_prefix'], "TEST1")
        self.assertEqual(req_module.config['root_module'], True)
        self.assertEqual(req_module.config['req_number_format'], "time_hash")

        self.assertEqual(req_module.modules['test_init2'].config['module_prefix'], "TEST2")
        self.assertEqual(req_module.modules['test_init2'].config['root_module'], True)
        self.assertEqual(req_module.modules['test_init2'].config['req_number_format'], "time_hash")

        self.assertEqual(
            req_module.modules['test_init2'].modules['test_init3'].config['module_prefix'], "TEST3")
        self.assertEqual(
            req_module.modules['test_init2'].modules['test_init3'].config['root_module'], False)
        self.assertEqual(
            req_module.modules['test_init2'].modules['test_init3'].config['req_number_format'],
            "time_hash")

        self.assertEqual(
            req_module.modules['test_init2'].modules['test_init4'].config['module_prefix'], "TEST4")
        self.assertEqual(
            req_module.modules['test_init2'].modules['test_init4'].config['root_module'], False)
        self.assertEqual(req_module.modules['test_init2'].modules['test_init4'].config['req_number_format'],
            "numbers")

        shutil.rmtree("test_init", ignore_errors=True)

    def test_export(self):
        if os.path.exists("test_export"):
            shutil.rmtree("test_export", ignore_errors=True)

        os.system('../git-reqs init --module test_export --module_prefix TEST1')
        os.system('cd test_export && ../../git-reqs init --module test_export2 --module_prefix TEST2 --root_module')

        req_module = requirementmodule.requirement_module("./test_export")

        # Create some reqs
        req_1 = {'Req-Id': '', 'Type': 'Requirement', 'Description': 'Req desc 1',
                 'upward_links': 'extern:Ext_1'}
        req_2 = {'Req-Id': '', 'Type': 'Requirement', 'Description': 'Req desc 2',
                 'downward_links': 'verifies:Test_1', 'upward_links': 'refines:Req_1'}
        test_1 = {'Req-Id': '', 'Type': 'Testcase', 'Description': 'Test desc 1'}

        req_1['Req-Id'] = req_module.add_req(req_1)
        req_2['Req-Id'] = req_module.add_req(req_2)
        test_1['Req-Id'] = req_module.modules['test_export2'].add_req(test_1)

        for field in req_1.keys():
            self.assertTrue(field in req_module.fields)
        for field in req_2.keys():
            self.assertTrue(field in req_module.fields)
        for field in test_1.keys():
            self.assertTrue(field in req_module.modules['test_export2'].fields)

        req_module.write_reqs()

        os.system(('../git-reqs export --project_root ./test_export --format md'))
        os.system(('cd test_export && ../../git-reqs export --format xls'))
        os.system(('cd test_export && ../../git-reqs export --module test_export2 --format xls'))

        wb1 = load_workbook('./test_export/TEST1.xlsx')


        sheet = wb1['TEST1']
        for col, field in enumerate(req_module.fields):
            self.assertEqual(sheet.cell(1, col+1).value, field)
            if field in req_1.keys():
                self.assertEqual(sheet.cell(2, col+1).value, req_1[field])
            if field in req_2.keys():
                self.assertEqual(sheet.cell(3, col+1).value, req_2[field])

        wb2 = load_workbook('./test_export/test_export2/TEST1_TEST2.xlsx')
        sheet = wb2['TEST1_TEST2']
        for col, field in enumerate(req_module.modules['test_export2'].fields):
            self.assertEqual(sheet.cell(1, col+1).value, field)

            if field in test_1.keys():
                self.assertEqual(sheet.cell(2, col+1).value, test_1[field])

        shutil.rmtree("test_export", ignore_errors=True)

    def test_import(self):
        if os.path.exists("test_import"):
            shutil.rmtree("test_import", ignore_errors=True)

        os.system('../git-reqs init --module test_import --module_prefix TEST1 --req_numbering_format numbers')
        os.system('cd test_import && ../../git-reqs init --module test_import2 --module_prefix TEST2 --root_module --req_numbering_format numbers')

        wb = Workbook()

        # Create some reqs
        req_1 = {'Req-Id': '', 'Type': 'Requirement', 'Description': 'Req desc 1',
                 'upward_links': 'extern:Ext_1'}
        req_2 = {'Req-Id': '', 'Type': 'Requirement', 'Description': 'Req desc 2',
                 'downward_links': 'verifies:TEST1_TEST2_1', 'upward_links': 'refines:TEST1_1'}
        test_1 = {'Req-Id': '', 'Type': 'Testcase', 'Description': 'Test desc 1'}

        fields = list(req_1.keys())
        fields.extend(list(req_2.keys()))
        fields.extend(list(test_1.keys()))
        fields = set(fields)

        sheet1 = wb[wb.sheetnames[0]]
        sheet2 = wb.copy_worksheet(wb[wb.sheetnames[0]])
        sheet1.title = "TEST1"
        sheet2.title = "TEST1_TEST2"

        for col, field in enumerate(fields):
            sheet1.cell(1, col + 1, field)
            sheet2.cell(1, col + 1, field)
            if field in req_1.keys():
                sheet1.cell(2, col + 1, req_1[field])
            if field in req_2.keys():
                sheet1.cell(3, col + 1, req_2[field])
            if field in test_1.keys():
                sheet2.cell(2, col + 1, test_1[field])

        wb.save("./test_import/TEST1.xlsx")

        os.system(('cd test_import && ../../git-reqs import --format xlsx'))


        req_module = requirementmodule.requirement_module("./test_import")

        for field in fields:
            if field != 'Req-Id':
                if field in req_1.keys():
                    self.assertEqual(req_module.reqs.nodes["TEST1_1"][field], req_1[field])
                if field in req_2.keys():
                    self.assertEqual(req_module.reqs.nodes["TEST1_2"][field], req_2[field])
                if field in test_1.keys():
                    self.assertEqual(req_module.modules["test_import2"].reqs.nodes["TEST1_TEST2_1"][field], test_1[field])

        os.system(('../git-reqs import --project_root ./test_import --format xlsx'))
        os.system(('cd test_import && ../../git-reqs import --format xls --file TEST1.xlsx'))

        shutil.rmtree("test_import", ignore_errors=True)

    def test_report(self):
        if os.path.exists("test_report"):
            shutil.rmtree("test_report", ignore_errors=True)

        os.system('../git-reqs init --module test_report --module_prefix TEST1')
        os.system('cd test_report && ../../git-reqs init --module test_report2 --module_prefix TEST2 --root_module')

        req_module = requirementmodule.requirement_module("./test_report")

        # Create some reqs
        test_1 = {'Req-Id': '', 'Type': 'Testcase', 'Description': 'Test desc 1'}
        test_1['Req-Id'] = req_module.modules['test_report2'].add_req(test_1)
        req_1 = {'Req-Id': '', 'Type': 'Requirement', 'Description': 'Req desc 1',
                 'upward_links': 'extern:Ext_1'}
        req_1['Req-Id'] = req_module.add_req(req_1)
        req_2 = {'Req-Id': '', 'Type': 'Requirement', 'Description': 'Req desc 2',
                 'downward_links': 'verifies:%s' % test_1['Req-Id'], 'upward_links': 'refines:%s' % req_1['Req-Id']}
        req_2['Req-Id'] = req_module.add_req(req_2)

        req_module.write_reqs()

        os.system(('../git-reqs report --project_root ./test_report --type test_coverage --dont_show_output'))
        os.system(('../git-reqs report --project_root ./test_report --type relations --dont_show_output'))

        shutil.rmtree("test_report", ignore_errors=True)
        
    def test_links(self):
        if os.path.exists("test_links"):
            shutil.rmtree("test_links", ignore_errors=True)

        os.system('../git-reqs init --module test_links --module_prefix ROOT --root_module')
        os.system('cd test_links && ../../git-reqs init --module ProductSpec --module_prefix PROD')
        os.system('cd test_links && ../../git-reqs init --module SubSystemRoot --module_prefix SubSystem')
        os.system('cd test_links/SubSystemRoot && ../../../git-reqs init --module SystemRS --module_prefix SystemRS')
        os.system('cd test_links/SubSystemRoot && ../../../git-reqs init --module Test --module_prefix Test')

        req_module = requirementmodule.requirement_module("./test_links")
        
        # Create some reqs
        ProdReq_1 = {'Req-Id': '', 'Type': 'Requirement', 'Description': 'Prod desc 1'}
        ProdReq_1['Req-Id'] = req_module.modules['ProductSpec'].add_req(ProdReq_1).split('_')[-1]

        # Link a subsystem requirement to an external requirement
        SubSystemReq1 = {'Req-Id': '', 'Type': 'Requirement', 'Description': 'Req desc 1',
                 'upward_links': 'extern:Ext_1'}        
        SubSystemReq1['Req-Id'] = req_module.modules['SubSystemRoot'].modules['SystemRS'].add_req(SubSystemReq1).split('_')[-1]

        # Link a test-case to a requirement on a sibling level module
        SubSystemTest1 = {'Req-Id': '', 'Type': 'Testcase', 'Description': 'Req desc 2',
                 'upward_links': 'verfies:SystemRS_%s' % SubSystemReq1['Req-Id']}
        SubSystemTest1['Req-Id'] = req_module.modules['SubSystemRoot'].modules['Test'].add_req(SubSystemTest1).split('_')[-1]

        # Link a requirement to a requirement in the same module, using only the number, and to the two-down-one-up
        # Production spec.
        SubSystemReq2 = {'Req-Id': '', 'Type': 'Requirement', 'Description': 'Req desc 1',
                         'upward_links': 'refines:%s,refines:PROD_%s' % (SubSystemReq1['Req-Id'], ProdReq_1['Req-Id'])}
        SubSystemReq2['Req-Id'] = req_module.modules['SubSystemRoot'].modules['SystemRS'].add_req(SubSystemReq2).split('_')[-1]

        ProdReq_2 = {'Req-Id': '', 'Type': 'Requirement', 'Description': 'Prod desc 2',
                     'downward_links': 'refines:SubSystem_SystemRS_%s' % SubSystemReq1['Req-Id']}
        ProdReq_2['Req-Id'] = req_module.modules['ProductSpec'].add_req(ProdReq_2).split('_')[-1]

        # Write the reqs and load the module again to get the links read properly
        req_module.write_reqs()
        req_module = requirementmodule.requirement_module("./test_links")


        self.assertTrue(('Ext_1', 'ROOT_SubSystem_SystemRS_%s' % SubSystemReq1['Req-Id'])\
                        in list(req_module.reqs.edges))
        self.assertTrue(('ROOT_SubSystem_SystemRS_%s' % SubSystemReq1['Req-Id'],
                         'ROOT_SubSystem_Test_%s' % SubSystemTest1['Req-Id']) \
                        in list(req_module.reqs.edges))
        self.assertTrue(('ROOT_PROD_%s' % ProdReq_1['Req-Id'], 'ROOT_SubSystem_SystemRS_%s' % SubSystemReq2['Req-Id']) \
                        in list(req_module.reqs.edges))
        self.assertTrue(('ROOT_PROD_%s' % ProdReq_2['Req-Id'],
                         'ROOT_SubSystem_SystemRS_%s' % SubSystemReq1['Req-Id']) \
                        in list(req_module.reqs.edges))

        shutil.rmtree("test_links", ignore_errors=True)
        


if __name__ == '__main__':
    unittest.main()




