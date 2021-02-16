from openpyxl import load_workbook
import os
import yaml
from pathlib import Path
import re
import networkx as nx

def import_from_xls(reqmodule, req_xls, wb=None):
    if not wb:
        wb = load_workbook(req_xls)

    if reqmodule.module_prefix in wb.sheetnames:
        sheet = wb[reqmodule.module_prefix]
    else:
        sheet = None

    if sheet:
        print(reqmodule.module_prefix)
        reqmodule.clear_ordered_req_list()

        fields = [field.value for field in list(sheet.iter_rows(1, 1))[0]]
        req = {}
        for row in range(2, sheet.max_row+1):
            for col, field in enumerate(fields):
                # Don't consider columns with empty field name,
                # empty fields
                # nor the 'non_stored_fields' columns
                if field and field != 'non_stored_fields':
                    req[field] = sheet.cell(row, col+1).value
                    if not req[field] or req[field] == 'None':
                        req[field] = ''

            # Check that the requirement is not empty before adding
            if [f for f in req.values() if f]:
                reqmodule.add_req(req)

    for submodule in reqmodule.modules.values():
        import_from_xls(submodule, req_xls, wb=wb)


def import_from_markdown(reqmodule, req_md):
    req = {}
    record_req = False
    with open(req_md, 'r') as md_file:
        md_line = md_file.readline()
        while md_line:
            if md_line[0:6] == '<!--- ' and md_line[6:13] == 'gitreq:':
                if md_line[13:19] == 'start:':
                    req = {}
                    req['Req-Id'] = md_line[19:].replace('-->\n', '')
                    req['Description'] = ""
                    record_req = True
                elif md_line[13:18] == 'stop:':
                    req['Description'] = req['Description'].replace(
                        '[' + req['Req-Id'] + ']', '').rstrip()
                    reqmodule.add_req(req)
                    record_req = False
            elif record_req:
                req['Description'] = req['Description'] + md_line
            md_line = md_file.readline()

def check_for_req_links_and_update(reqmodule, string, string_changed):
    pattern = reqmodule.get_link_regex()
    res = re.search(pattern, string)
    if res:
        dest_req_id = reqmodule.module_prefix + '_' + res.groups()[2].replace("\"", "")
        dest_req = reqmodule.reqs.nodes[dest_req_id]

        link_name = res.groups()[0].split(':')
        # Create a new req/test if required
        if link_name[0].startswith("?"):
            req_type = link_name[1]
            desc = link_name[2].replace('..', ' ') if len(link_name) > 2 else ""

            name = reqmodule.add_req_with_path(link_name[0][1:], {"Req-Id": "",
                                      "Type": req_type,
                                      "Description": desc})

            string = string.replace(res.groups()[0], name)
            string_changed = True
        else:
            name = res.groups()[0].split(':')[0]

        link = ':'.join([res.groups()[1], name])
        if not link in dest_req['downward_links']:
            if not dest_req['downward_links'] == "":
                dest_req['downward_links'] += ','
            dest_req['downward_links'] += link

    return string, string_changed



def parse_requirement_links(reqmodule, source_root):
    if reqmodule.config['req_version'] >= 0.3:
        for src_pth in reqmodule.config['source_paths']:
            for ext in reqmodule.config['source_extensions']:
                files = Path(source_root + '/' + src_pth).rglob('*.' + ext)
                for f in files:
                    print('Checking source file: %s' % f.name)
                    with open(f.absolute(), 'r') as src_file:
                        with open(str(f.absolute()) + "_tmp", 'w') as new_src_file:
                            file_changed = False
                            for line_nr, line in enumerate(src_file):
                                line, file_changed = check_for_req_links_and_update(reqmodule, line, file_changed)
                                new_src_file.write(line)

                    if file_changed:
                        os.remove(f.absolute())
                        os.rename(str(f.absolute()) + "_tmp", f.absolute())
                    else:
                        os.remove(str(f.absolute()) + "_tmp")
        reqmodule.write_reqs()

    else:
        print('Requirement link parsing is only available from req version 0.3 and up')


def add_test_results(module_path, test_results_file):
    if os.path.exists(module_path + '/test_results.temp.yaml'):
        with open(module_path + '/test_results.temp.yaml', 'r') as testlist_file:
            test_result_files = yaml.safe_load(testlist_file)
            test_result_files.append(test_results_file)
    else:
        test_result_files = [test_results_file]

    with open(module_path + '/test_results.temp.yaml', 'w') as testlist_file:
        yaml.dump(test_result_files, testlist_file)
